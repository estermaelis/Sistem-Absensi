import csv
import os
from io import BytesIO, StringIO
from datetime import datetime
from src.database import get_db_connection
from mysql.connector import Error

# Column headers shared across all export formats
EXPORT_HEADERS = ['NIP', 'Nama', 'Departemen', 'Jenis Kelamin', 'Tanggal',
                  'Jam Masuk', 'Jam Pulang', 'Status', 'Confidence']


def fetch_attendance_rows(start_date=None, end_date=None):
    """Query attendance records, optionally filtered by date range.
    Returns a list of dict rows (may be empty). Raises on DB failure."""
    connection = get_db_connection()
    if connection is None:
        raise RuntimeError("Gagal terhubung ke database")

    base_query = """
        SELECT
            e.nip,
            e.name,
            d.name AS department_name,
            e.gender,
            a.attendance_date,
            a.check_in_time,
            a.check_out_time,
            a.status,
            a.confidence
        FROM attendance a
        INNER JOIN employees e ON a.employee_id = e.id
        LEFT JOIN departments d ON e.department_id = d.id
    """

    cursor = connection.cursor(dictionary=True)
    try:
        if start_date and end_date:
            cursor.execute(base_query + " WHERE a.attendance_date BETWEEN %s AND %s"
                           " ORDER BY a.attendance_date DESC, a.check_in_time ASC",
                           (start_date, end_date))
        elif start_date:
            cursor.execute(base_query + " WHERE a.attendance_date >= %s"
                           " ORDER BY a.attendance_date DESC, a.check_in_time ASC",
                           (start_date,))
        else:
            cursor.execute(base_query +
                           " ORDER BY a.attendance_date DESC, a.check_in_time ASC")
        rows = cursor.fetchall()
    finally:
        cursor.close()
        connection.close()
    return rows


def _row_to_values(row):
    """Convert a DB row dict into a list of display strings matching EXPORT_HEADERS."""
    return [
        row['nip'],
        row['name'],
        row['department_name'] or '',
        'Laki-laki' if row['gender'] == 'L' else 'Perempuan',
        row['attendance_date'].strftime('%Y-%m-%d') if row['attendance_date'] else '',
        str(row['check_in_time']) if row['check_in_time'] else '',
        str(row['check_out_time']) if row['check_out_time'] else '',
        row['status'],
        f"{row['confidence']:.2f}" if row['confidence'] is not None else '',
    ]


def build_csv(rows):
    """Build a CSV file from rows. Returns BytesIO."""
    text_buffer = StringIO()
    writer = csv.writer(text_buffer)
    writer.writerow(EXPORT_HEADERS)
    for row in rows:
        writer.writerow(_row_to_values(row))
    return BytesIO(text_buffer.getvalue().encode('utf-8-sig'))


def build_excel(rows):
    """Build an XLSX file from rows. Returns BytesIO."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Absensi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    ws.append(EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append(_row_to_values(row))

    # Auto-size columns based on content
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        max_len = len(header)
        for row in rows:
            value = str(_row_to_values(row)[col_idx - 1])
            max_len = max(max_len, len(value))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_pdf(rows, title="Laporan Absensi Karyawan"):
    """Build a PDF file from rows. Returns BytesIO."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']),
                Paragraph(f"Dibuat: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']),
                Spacer(1, 0.5*cm)]

    data = [EXPORT_HEADERS] + [_row_to_values(row) for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output


def export_report():
    """CLI: Export attendance report to CSV"""
    print("=" * 50)
    print("EXPORT LAPORAN ABSENSI")
    print("=" * 50)

    print("\nMasukkan rentang tanggal (kosongkan untuk semua data)")
    start_date = input("Tanggal awal (YYYY-MM-DD): ").strip() or None
    end_date = input("Tanggal akhir (YYYY-MM-DD): ").strip() or None

    try:
        rows = fetch_attendance_rows(start_date, end_date)
    except Exception as e:
        print(f"Error: {e}")
        return

    if not rows:
        print("\nTidak ada data absensi yang ditemukan!")
        return

    print(f"\nDitemukan {len(rows)} record absensi.")

    os.makedirs('exports', exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"laporan_absensi_{timestamp}.csv"
    filepath = os.path.join('exports', filename)

    with open(filepath, 'wb') as f:
        f.write(build_csv(rows).getvalue())

    print(f"\nLaporan berhasil diekspor ke: {filepath}")
    print(f"Total record: {len(rows)}")


if __name__ == "__main__":
    export_report()
